import pandas as pd
import os
import re
from typing import List, Dict, Set, Any, Optional, Tuple
from datetime import datetime
from dataclasses import dataclass
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

@dataclass
class TestCase:
    """Individual test case data structure"""
    tc_id: str
    field_name: str
    category: str
    type_of_validation: str
    test_objective: str
    request_response_field: str
    test_steps: str
    expected_result: str
    mapping_correlation: str
    manual_automation: str
    status: str = "pending"  # pending, approved, rejected
    timestamp: str = ""
    original_content: str = ""  # Store original parsed content

class FieldSession:
    """Session data for a single field"""
    def __init__(self, field_name: str):
        self.field_name = field_name
        self.test_cases: Dict[str, TestCase] = {}
        self.next_tc_number = 1
        self.creation_time = datetime.now()
        self.completion_time = None
        self.status = "in_progress"  # in_progress, completed
        self.approved_count = 0
        self.rejected_count = 0
        self.total_generated = 0

class TestCaseManager:
    """
    Enhanced test case manager with multi-field support and sequential TC ID management
    """
    
    def __init__(self):
        # Multi-field storage
        self.field_sessions: Dict[str, FieldSession] = {}
        self.global_tc_counter = 1  # Global counter for sequential IDs across all fields
        
        # Error tracking
        self.parse_errors = []
        
        # Session metadata
        self.session_start_time = datetime.now()
        self.current_field = None
    
    def _normalize_validation_type(self, val_type: str) -> str:
        """Normalize validation type to standard values"""
        if not val_type:
            return "Field Validation - Positive"
        
        val_type_lower = val_type.lower()
        if "positive" in val_type_lower and "field" in val_type_lower:
            return "Field Validation - Positive"
        elif "negative" in val_type_lower and "field" in val_type_lower:
            return "Field Validation - Negative"
        elif "positive" in val_type_lower and "business" in val_type_lower:
            return "Business Validation - Positive"
        elif "negative" in val_type_lower and "business" in val_type_lower:
            return "Business Validation - Negative"
        else:
            return "Field Validation - Positive"
    
    def _determine_automation_mode(self, val_type: str, provided_mode: str) -> str:
        """Determine automation mode based on validation type"""
        if provided_mode and provided_mode.strip():
            return provided_mode.strip()
        
        if "Business" in val_type:
            return "Manual"
        else:
            return "Automation"
    
    def _ensure_field_session(self, field_name: str) -> FieldSession:
        """Ensure field session exists, create if needed"""
        if field_name not in self.field_sessions:
            self.field_sessions[field_name] = FieldSession(field_name)
            print(f"[INFO] Created new field session for: {field_name}")
        
        return self.field_sessions[field_name]
    
    def parse_and_add_test_cases(self, raw_text: str, default_mapping: str = "", 
                                field_name: str = None) -> List[str]:
        """
        Parse AI response and add test cases to appropriate field session
        Returns list of new TC IDs created
        """
        
        if not raw_text or not raw_text.strip():
            print("[WARN] Empty input provided for parsing")
            return []
        
        if not field_name:
            print("[ERROR] Field name is required for multi-field support")
            return []
        
        # Ensure field session exists
        field_session = self._ensure_field_session(field_name)
        self.current_field = field_name
        
        new_tc_ids = []
        lines = raw_text.strip().splitlines()
        
        print(f"[INFO] Parsing test cases for field: {field_name}")
        
        for line_num, line in enumerate(lines, 1):
            line = line.strip()
            
            # Skip empty lines, headers, and obvious non-data lines
            if (not line or 
                line.startswith('Category') or 
                line.startswith('---') or 
                line.startswith('===') or
                'Test Case ID' in line or
                line.startswith('Sorry') or
                line.startswith('I don\'t') or
                len(line) < 10):
                continue
            
            try:
                # Try tab separation first
                parts = line.split('\t')
                
                # Fallback to pipe separation if not enough parts
                if len(parts) < 6:
                    parts = line.split('|')
                
                if len(parts) < 6:
                    error_msg = f"Line {line_num}: Insufficient columns ({len(parts)}) - '{line[:50]}...'"
                    self.parse_errors.append(error_msg)
                    print(f"[WARN] {error_msg}")
                    continue
                
                # Clean and pad parts
                parts = [p.strip() for p in parts]
                while len(parts) < 9:
                    parts.append("")
                
                # Extract and validate components
                category = parts[0] or "Functional"
                tc_id_provided = parts[1].strip()  # Usually blank
                val_type = self._normalize_validation_type(parts[2])
                objective = parts[3]
                req_field = parts[4] or "Request"
                steps = parts[5]
                expected = parts[6] or "Expected result"
                mapping = parts[7] or default_mapping
                mode = self._determine_automation_mode(val_type, parts[8])
                
                # Validate required fields
                if not objective or len(objective.strip()) < 5:
                    error_msg = f"Line {line_num}: Missing or invalid test objective"
                    self.parse_errors.append(error_msg)
                    continue
                
                if not steps or len(steps.strip()) < 5:
                    error_msg = f"Line {line_num}: Missing or invalid test steps"
                    self.parse_errors.append(error_msg)
                    continue
                
                # Generate sequential TC ID (global counter)
                tc_id = f"TC_{self.global_tc_counter:03d}"
                self.global_tc_counter += 1
                
                # Create test case object
                test_case = TestCase(
                    tc_id=tc_id,
                    field_name=field_name,
                    category=category,
                    type_of_validation=val_type,
                    test_objective=objective,
                    request_response_field=req_field,
                    test_steps=steps,
                    expected_result=expected,
                    mapping_correlation=mapping,
                    manual_automation=mode,
                    status="pending",
                    timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    original_content=line
                )
                
                # Add to field session
                field_session.test_cases[tc_id] = test_case
                field_session.total_generated += 1
                new_tc_ids.append(tc_id)
                
                print(f"[DEBUG] Created {tc_id} for field {field_name}: {objective[:50]}...")
                
            except Exception as e:
                error_msg = f"Line {line_num}: Parse error - {str(e)}"
                self.parse_errors.append(error_msg)
                print(f"[ERROR] {error_msg}")
                continue
        
        added_count = len(new_tc_ids)
        if added_count > 0:
            print(f"[INFO] Successfully parsed {added_count} test cases for field {field_name}")
        else:
            print(f"[WARN] No valid test cases found in input for field {field_name}")
        
        return new_tc_ids
    
    def approve_test_cases(self, tc_ids: List[str]) -> Dict[str, List[str]]:
        """Approve specific test cases across all fields"""
        approved = []
        not_found = []
        
        for tc_id in tc_ids:
            test_case = self._find_test_case_by_id(tc_id)
            if test_case:
                test_case.status = "approved"
                
                # Update field session stats
                field_session = self.field_sessions[test_case.field_name]
                field_session.approved_count += 1
                
                approved.append(tc_id)
                print(f"[DEBUG] Approved {tc_id} from field {test_case.field_name}")
            else:
                not_found.append(tc_id)
        
        return {"approved": approved, "not_found": not_found}
    
    def reject_test_cases(self, tc_ids: List[str]) -> Dict[str, List[str]]:
        """Reject specific test cases (mark as rejected, don't delete)"""
        rejected = []
        not_found = []
        
        for tc_id in tc_ids:
            test_case = self._find_test_case_by_id(tc_id)
            if test_case:
                test_case.status = "rejected"
                
                # Update field session stats
                field_session = self.field_sessions[test_case.field_name]
                field_session.rejected_count += 1
                
                rejected.append(tc_id)
                print(f"[DEBUG] Rejected {tc_id} from field {test_case.field_name}")
            else:
                not_found.append(tc_id)
        
        return {"rejected": rejected, "not_found": not_found}
    
    def _find_test_case_by_id(self, tc_id: str) -> Optional[TestCase]:
        """Find test case by ID across all field sessions"""
        for field_session in self.field_sessions.values():
            if tc_id in field_session.test_cases:
                return field_session.test_cases[tc_id]
        return None
    
    def get_test_case_by_id(self, tc_id: str) -> Optional[Dict[str, Any]]:
        """Get test case data by ID (for external use)"""
        test_case = self._find_test_case_by_id(tc_id)
        if test_case:
            return self._test_case_to_dict(test_case)
        return None
    
    def get_field_test_cases(self, field_name: str) -> List[Dict[str, Any]]:
        """Get all test cases for a specific field"""
        if field_name not in self.field_sessions:
            return []
        
        field_session = self.field_sessions[field_name]
        return [self._test_case_to_dict(tc) for tc in field_session.test_cases.values()]
    
    def get_field_approved_cases(self, field_name: str) -> List[Dict[str, Any]]:
        """Get approved test cases for a specific field"""
        if field_name not in self.field_sessions:
            return []
        
        field_session = self.field_sessions[field_name]
        approved_cases = [tc for tc in field_session.test_cases.values() if tc.status == "approved"]
        return [self._test_case_to_dict(tc) for tc in approved_cases]
    
    def get_cases_by_status(self, status: str) -> List[Dict[str, Any]]:
        """Get test cases by status across all fields"""
        cases = []
        for field_session in self.field_sessions.values():
            field_cases = [tc for tc in field_session.test_cases.values() if tc.status == status]
            cases.extend([self._test_case_to_dict(tc) for tc in field_cases])
        return cases
    
    def get_all_cases(self) -> List[Dict[str, Any]]:
        """Get all test cases across all fields"""
        all_cases = []
        for field_session in self.field_sessions.values():
            all_cases.extend([self._test_case_to_dict(tc) for tc in field_session.test_cases.values()])
        
        # Sort by TC ID for consistent ordering
        all_cases.sort(key=lambda x: x.get('Test Case ID', ''))
        return all_cases
    
    def get_approved_cases(self) -> List[Dict[str, Any]]:
        """Get all approved test cases across all fields"""
        return self.get_cases_by_status("approved")
    
    def get_pending_cases(self) -> List[Dict[str, Any]]:
        """Get all pending test cases across all fields"""
        return self.get_cases_by_status("pending")
    
    def _test_case_to_dict(self, test_case: TestCase) -> Dict[str, Any]:
        """Convert TestCase object to dictionary for external use"""
        return {
            "Category": test_case.category,
            "Test Case ID": test_case.tc_id,
            "Type of Validation": test_case.type_of_validation,
            "Test Objective": test_case.test_objective,
            "Request/Response Field": test_case.request_response_field,
            "Test Steps": test_case.test_steps,
            "Expected Result": test_case.expected_result,
            "Mapping Correlation": test_case.mapping_correlation,
            "Manual/Automation": test_case.manual_automation,
            "Status": test_case.status,
            "Field Name": test_case.field_name,
            "Timestamp": test_case.timestamp
        }
    
    def complete_field(self, field_name: str) -> Dict[str, Any]:
        """Mark a field as completed and return summary"""
        if field_name not in self.field_sessions:
            return {"error": f"Field {field_name} not found"}
        
        field_session = self.field_sessions[field_name]
        field_session.status = "completed"
        field_session.completion_time = datetime.now()
        
        approved_cases = [tc for tc in field_session.test_cases.values() if tc.status == "approved"]
        
        print(f"[INFO] Completed field: {field_name} with {len(approved_cases)} approved cases")
        
        return {
            "field_name": field_name,
            "approved_count": len(approved_cases),
            "rejected_count": field_session.rejected_count,
            "total_generated": field_session.total_generated,
            "completion_time": field_session.completion_time,
            "status": "completed"
        }
    
    def get_multi_field_stats(self) -> Dict[str, Any]:
        """Get comprehensive statistics across all fields"""
        
        total_approved = 0
        total_rejected = 0
        total_pending = 0
        total_generated = 0
        completed_fields = 0
        
        field_breakdown = {}
        
        for field_name, field_session in self.field_sessions.items():
            approved = len([tc for tc in field_session.test_cases.values() if tc.status == "approved"])
            rejected = len([tc for tc in field_session.test_cases.values() if tc.status == "rejected"])
            pending = len([tc for tc in field_session.test_cases.values() if tc.status == "pending"])
            
            total_approved += approved
            total_rejected += rejected
            total_pending += pending
            total_generated += field_session.total_generated
            
            if field_session.status == "completed":
                completed_fields += 1
            
            field_breakdown[field_name] = {
                "approved": approved,
                "rejected": rejected,
                "pending": pending,
                "total": field_session.total_generated,
                "status": field_session.status
            }
        
        return {
            "total_approved": total_approved,
            "total_rejected": total_rejected,
            "total_pending": total_pending,
            "total_generated": total_generated,
            "total_fields": len(self.field_sessions),
            "completed_fields": completed_fields,
            "field_breakdown": field_breakdown,
            "session_duration": str(datetime.now() - self.session_start_time)
        }
    
    def get_status_summary(self) -> Dict[str, int]:
        """Get status summary across all fields"""
        stats = self.get_multi_field_stats()
        return {
            "pending": stats["total_pending"],
            "approved": stats["total_approved"],
            "rejected": stats["total_rejected"]
        }
    
    def export_multi_field_session(self, output_file: str) -> bool:
        """
        Export all approved test cases from all completed fields to single Excel
        with sequential TC IDs and field separation
        """
        
        # Get all approved cases from completed fields
        approved_cases_by_field = {}
        total_approved = 0
        
        for field_name, field_session in self.field_sessions.items():
            if field_session.status == "completed":
                approved_cases = [tc for tc in field_session.test_cases.values() if tc.status == "approved"]
                if approved_cases:
                    approved_cases_by_field[field_name] = approved_cases
                    total_approved += len(approved_cases)
        
        if total_approved == 0:
            print("[ERROR] No approved test cases from completed fields to export")
            return False
        
        print(f"[INFO] Exporting {total_approved} approved test cases from {len(approved_cases_by_field)} completed fields")
        
        try:
            return self._create_multi_field_excel(approved_cases_by_field, output_file)
            
        except Exception as e:
            print(f"[ERROR] Export failed: {str(e)}")
            return False
    
    def _create_multi_field_excel(self, approved_cases_by_field: Dict[str, List[TestCase]], 
                                 output_file: str) -> bool:
        """Create Excel file with multi-field test cases and sequential TC IDs"""
        
        # Prepare data with field separators and sequential TC IDs
        export_data = []
        tc_counter = 1
        
        for field_name, test_cases in approved_cases_by_field.items():
            # Sort test cases by original TC ID to maintain some order
            test_cases.sort(key=lambda x: x.tc_id)
            
            # Add field separator row
            separator_row = {
                "Category": f"=== {field_name.upper()} TEST CASES ===",
                "Test Case ID": "",
                "Type of Validation": "",
                "Test Objective": "",
                "Request/Response Field": "",
                "Test Steps": "",
                "Expected Result": "",
                "Mapping Correlation": "",
                "Manual/Automation": ""
            }
            export_data.append(separator_row)
            
            # Add test cases with sequential TC IDs
            for test_case in test_cases:
                case_dict = {
                    "Category": test_case.category,
                    "Test Case ID": f"TC_{tc_counter:03d}",  # Sequential across all fields
                    "Type of Validation": test_case.type_of_validation,
                    "Test Objective": test_case.test_objective,
                    "Request/Response Field": test_case.request_response_field,
                    "Test Steps": test_case.test_steps,
                    "Expected Result": test_case.expected_result,
                    "Mapping Correlation": test_case.mapping_correlation,
                    "Manual/Automation": test_case.manual_automation
                }
                export_data.append(case_dict)
                tc_counter += 1
            
            # Add blank separator row between fields
            blank_row = {col: "" for col in separator_row.keys()}
            export_data.append(blank_row)
        
        # Create DataFrame
        df = pd.DataFrame(export_data)
        
        # Create Excel with formatting
        self._create_formatted_excel(df, output_file, approved_cases_by_field)
        
        print(f"[INFO] Successfully exported multi-field session to: {output_file}")
        return True
    
    def _create_formatted_excel(self, df: pd.DataFrame, output_file: str, 
                               approved_cases_by_field: Dict[str, List[TestCase]]):
        """Create professionally formatted Excel file"""
        
        wb = Workbook()
        
        # Main test cases sheet
        ws = wb.active
        ws.title = "Test Cases"
        
        # Define styles
        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        separator_font = Font(name='Calibri', size=11, bold=True, color='000000')
        separator_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
        
        data_font = Font(name='Calibri', size=10)
        data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        # Add headers
        for col, header in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Add data with formatting
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
                cell.border = thin_border
                
                # Format field separator rows
                if str(value).startswith("===") and str(value).endswith("==="):
                    cell.font = separator_font
                    cell.fill = separator_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center', bold=True)
                else:
                    cell.font = data_font
                    cell.alignment = data_alignment
                
                # Color coding for validation types
                if col_idx == 3 and value:  # Type of Validation column
                    if 'Positive' in str(value):
                        cell.fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
                    elif 'Negative' in str(value):
                        cell.fill = PatternFill(start_color='FFF0F0', end_color='FFF0F0', fill_type='solid')
                
                # Highlight manual tests
                if col_idx == 9 and str(value).lower() == 'manual':
                    cell.fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')
        
        # Auto-adjust column widths
        column_widths = [12, 15, 28, 50, 18, 50, 40, 35, 15]
        for col, width in enumerate(column_widths, 1):
            if col <= len(df.columns):
                column_letter = ws.cell(row=1, column=col).column_letter
                ws.column_dimensions[column_letter].width = width
        
        # Set row heights
        ws.row_dimensions[1].height = 35
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Create summary sheet
        summary_ws = wb.create_sheet("Session Summary")
        self._create_summary_sheet(summary_ws, approved_cases_by_field)
        
        # Save workbook
        wb.save(output_file)
    
    def _create_summary_sheet(self, ws, approved_cases_by_field: Dict[str, List[TestCase]]):
        """Create summary sheet with session statistics"""
        
        stats = self.get_multi_field_stats()
        
        # Title
        ws.cell(row=1, column=1, value="Multi-Field Test Case Session Summary").font = Font(size=14, bold=True)
        ws.cell(row=2, column=1, value=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Overall statistics
        ws.cell(row=4, column=1, value="Overall Statistics").font = Font(size=12, bold=True)
        ws.cell(row=5, column=1, value="Total Fields Processed:")
        ws.cell(row=5, column=2, value=stats["total_fields"])
        ws.cell(row=6, column=1, value="Completed Fields:")
        ws.cell(row=6, column=2, value=stats["completed_fields"])
        ws.cell(row=7, column=1, value="Total Test Cases Generated:")
        ws.cell(row=7, column=2, value=stats["total_generated"])
        ws.cell(row=8, column=1, value="Total Approved Cases:")
        ws.cell(row=8, column=2, value=stats["total_approved"])
        ws.cell(row=9, column=1, value="Session Duration:")
        ws.cell(row=9, column=2, value=stats["session_duration"])
        
        # Field breakdown
        ws.cell(row=11, column=1, value="Field Breakdown").font = Font(size=12, bold=True)
        ws.cell(row=12, column=1, value="Field Name").font = Font(bold=True)
        ws.cell(row=12, column=2, value="Approved").font = Font(bold=True)
        ws.cell(row=12, column=3, value="Total Generated").font = Font(bold=True)
        ws.cell(row=12, column=4, value="Status").font = Font(bold=True)
        
        row_num = 13
        for field_name, breakdown in stats["field_breakdown"].items():
            ws.cell(row=row_num, column=1, value=field_name)
            ws.cell(row=row_num, column=2, value=breakdown["approved"])
            ws.cell(row=row_num, column=3, value=breakdown["total"])
            ws.cell(row=row_num, column=4, value=breakdown["status"])
            row_num += 1
        
        # Auto-adjust column widths
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20
    
    def display_test_cases(self, cases: List[Dict[str, Any]], show_details: bool = True):
        """Display test cases in readable format (for console output)"""
        if not cases:
            print("No test cases to display.")
            return
        
        for case in cases:
            tc_id = case.get("Test Case ID", "N/A")
            status = case.get("Status", "pending")
            field_name = case.get("Field Name", "Unknown")
            val_type = case.get("Type of Validation", "N/A")
            objective = case.get("Test Objective", "N/A")
            
            # Status icon
            status_icons = {"approved": "✅", "rejected": "❌", "pending": "📝"}
            icon = status_icons.get(status, "📝")
            
            print(f"\n{icon} {tc_id} [{field_name}] - {val_type}")
            print(f"    📋 {objective}")
            
            if show_details:
                steps = case.get("Test Steps", "N/A")
                expected = case.get("Expected Result", "N/A")
                mode = case.get("Manual/Automation", "N/A")
                
                print(f"    🔧 Steps: {steps}")
                print(f"    ✅ Expected: {expected}")
                print(f"    🤖 Mode: {mode}")
    
    def clear_cases(self):
        """Clear all test cases and reset counters"""
        self.field_sessions.clear()
        self.global_tc_counter = 1
        self.parse_errors.clear()
        self.current_field = None
        self.session_start_time = datetime.now()
        print("[INFO] All test cases and field sessions cleared")
    
    def get_parse_errors(self) -> List[str]:
        """Get all parsing errors encountered"""
        return self.parse_errors.copy()
    
    # Legacy compatibility methods
    def export_to_excel(self, out_file: str) -> bool:
        """Legacy method - exports all approved cases"""
        return self.export_multi_field_session(out_file)