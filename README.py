# core/test_objective_core.py
import hashlib
import time
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List, Dict, Any, Optional, Tuple

from core.java_extractor import extract_java_code_blocks_with_cross_references, trim_code_context

class TestObjectiveGeneratorCore:
    def __init__(self, client, test_manager, src_dir: str):
        self.client = client
        self.test_manager = test_manager
        self.src_dir = src_dir
        self.failed_fields = []  # Track failed fields for reporting

    def _hash_text(self, text: str) -> str:
        return hashlib.sha256(text.encode("utf-8")).hexdigest()[:8]

    def _validate_field_data(self, field: dict) -> Tuple[bool, str]:
        """Validate field has minimum required data"""
        required_fields = ['field_name']
        
        for req_field in required_fields:
            if not field.get(req_field):
                return False, f"Missing required field: {req_field}"
        
        # Check for reasonable field name
        field_name = field.get('field_name', '').strip()
        if len(field_name) < 2:
            return False, f"Field name too short: '{field_name}'"
        
        # Validate backend_xpath if present
        backend_xpath = field.get('backend_xpath', '')
        if backend_xpath and not backend_xpath.replace('/', '').replace('_', '').replace('-', '').isalnum():
            return False, f"Invalid backend_xpath format: '{backend_xpath}'"
        
        return True, "Valid"

    def _format_fiservai_prompt(self, fields: List[dict], code_contexts: Dict[str, str]) -> str:
        """Format prompt for FiservAI's CONTEXT/QUESTION format"""
        
        # Build CONTEXT section
        context = """You are an expert QA test objectives generator for ESF APIs.

FIELD INFORMATION AND CODE CONTEXT:
"""
        
        for field in fields:
            fid = field.get("backend_xpath") or field.get("field_name", "unknown")
            fid_short = f"{field.get('field_name','F')}_{self._hash_text(fid)}"
            
            context += f"\n--- FIELD: {field.get('field_name', 'Unknown')} ---\n"
            for k, v in field.items():
                if v:  # Only include non-empty values
                    context += f"{k}: {v}\n"
            
            ctx = code_contexts.get(fid_short, "")
            if ctx:
                context += f"\nJAVA CODE CONTEXT:\n{ctx}\n"
            else:
                context += "\nNO JAVA CODE FOUND FOR THIS FIELD\n"
        
        # Build QUESTION section
        question = """Generate test cases in EXACTLY this format - 9 tab-separated columns:

Category | Test Case ID (leave blank) | Type of Validation | Test Objective | Request/Response Field | Test Steps | Expected Result | Mapping Correlation | Manual/Automation

REQUIREMENTS:
- Category: Always "Functional"
- Type of Validation: Must be one of: "Field Validation - Positive", "Field Validation - Negative", "Business Validation - Positive", "Business Validation - Negative"
- Manual/Automation: "Manual" for business validation, "Automation" for field validation
- Mapping Correlation: Use backend_xpath from field metadata

For each field, wrap output between markers:
===FIELD_START:<FIELD_ID>=== 
[test case rows here]
===FIELD_END:<FIELD_ID>===

Generate 2-4 test cases per field covering positive and negative scenarios based on the code context."""

        return f"====CONTEXT {context} ====QUESTION {question}"

    def _call_api_with_retry(self, prompt: str, max_retries: int = 3) -> Optional[str]:
        """Call API with retry logic and error handling"""
        
        for attempt in range(max_retries):
            try:
                print(f"[DEBUG] API call attempt {attempt + 1}/{max_retries}")
                response = self.client.chat_completion(prompt)
                content = response.choices[0].message.content.strip()
                
                # Check for FiservAI's "I don't know" responses
                if "sorry" in content.lower() and ("don't know" in content.lower() or "not sure" in content.lower()):
                    if attempt < max_retries - 1:
                        print(f"[WARN] AI responded with 'don't know', retrying with simpler prompt...")
                        # Create simpler fallback prompt
                        prompt = self._create_fallback_prompt(prompt)
                        time.sleep(2)  # Brief delay before retry
                        continue
                    else:
                        print(f"[ERROR] AI couldn't generate test cases after {max_retries} attempts")
                        return None
                
                return content
                
            except Exception as e:
                print(f"[ERROR] API call failed (attempt {attempt + 1}): {str(e)}")
                if attempt < max_retries - 1:
                    wait_time = 2 ** attempt  # Exponential backoff
                    print(f"[INFO] Waiting {wait_time} seconds before retry...")
                    time.sleep(wait_time)
                else:
                    print(f"[ERROR] All API retry attempts failed")
                    return None
        
        return None

    def _create_fallback_prompt(self, original_prompt: str) -> str:
        """Create simpler prompt when AI says 'I don't know'"""
        
        # Extract field info from original prompt
        if "FIELD INFORMATION" in original_prompt:
            context_part = original_prompt.split("====QUESTION")[0].replace("====CONTEXT ", "")
        else:
            context_part = "Generate basic test cases for API field validation."
        
        # Simpler question
        question = """Generate basic test cases in this simple format:

Functional	 	Field Validation - Positive	Test valid input	Request	Send valid data	Success expected	field/path	Automation
Functional	 	Field Validation - Negative	Test invalid input	Request	Send invalid data	Error expected	field/path	Automation

Just create 2 simple test cases following this exact pattern."""
        
        return f"====CONTEXT {context_part} ====QUESTION {question}"

    def generate_for_field(self, field: dict) -> bool:
        """Single-field generation (chatbot mode) with error handling"""
        
        # Validate field data
        is_valid, error_msg = self._validate_field_data(field)
        if not is_valid:
            print(f"[ERROR] Invalid field data: {error_msg}")
            return False
        
        try:
            backend_xpath = field.get("backend_xpath") or ""
            field_name = field.get("field_name", "")
            
            # Extract keywords safely
            keywords = []
            if backend_xpath:
                last_seg = backend_xpath.split("/")[-1]
                if last_seg:
                    keywords.append(last_seg)
            if field_name:
                keywords.append(field_name)
            
            if not keywords:
                print(f"[WARN] No valid keywords found for field, using generic search")
                keywords = ["validate", "check"]  # Fallback keywords
            
            print(f"[INFO] Extracting Java code for keywords: {keywords}")
            
            # Extract Java code with error handling
            try:
                snippets = extract_java_code_blocks_with_cross_references(
                    self.src_dir, keywords, max_depth=1
                )
                context = trim_code_context(snippets, max_chars=2000)
            except Exception as e:
                print(f"[WARN] Java extraction failed: {str(e)}, continuing without code context")
                context = ""
            
            fid_short = f"{field_name}_{self._hash_text(backend_xpath or field_name)}"
            
            # Generate prompt and call API
            prompt = self._format_fiservai_prompt([field], {fid_short: context})
            output = self._call_api_with_retry(prompt)
            
            if output is None:
                print(f"[ERROR] Failed to generate test cases for field: {field_name}")
                self.failed_fields.append(field_name)
                return False
            
            # Parse and store results
            success = self._parse_and_store(output, [field])
            if not success:
                print(f"[ERROR] Failed to parse test cases for field: {field_name}")
                self.failed_fields.append(field_name)
                return False
            
            return True
            
        except Exception as e:
            print(f"[ERROR] Unexpected error processing field {field.get('field_name', 'unknown')}: {str(e)}")
            self.failed_fields.append(field.get('field_name', 'unknown'))
            return False

    def bulk_generate(self, fields: List[dict], batch_size: int = 5, max_workers: int = 6) -> Dict[str, Any]:
        """Bulk generation with comprehensive error handling"""
        
        print(f"[INFO] Starting bulk generation for {len(fields)} fields")
        
        # Validate all fields first
        valid_fields = []
        invalid_fields = []
        
        for field in fields:
            is_valid, error_msg = self._validate_field_data(field)
            if is_valid:
                valid_fields.append(field)
            else:
                invalid_fields.append({
                    'field': field.get('field_name', 'unknown'),
                    'error': error_msg
                })
                print(f"[WARN] Skipping invalid field: {error_msg}")
        
        if not valid_fields:
            print("[ERROR] No valid fields to process")
            return {
                'success': False,
                'processed': 0,
                'failed': len(fields),
                'invalid_fields': invalid_fields,
                'error': 'No valid fields found'
            }
        
        print(f"[INFO] Processing {len(valid_fields)} valid fields in batches of {batch_size}")
        
        total_processed = 0
        total_failed = 0
        batch_failures = []
        
        # Process in batches
        for i in range(0, len(valid_fields), batch_size):
            batch_num = i // batch_size + 1
            total_batches = (len(valid_fields) + batch_size - 1) // batch_size
            batch = valid_fields[i:i + batch_size]
            
            print(f"[INFO] Processing batch {batch_num}/{total_batches} ({len(batch)} fields)")
            
            try:
                # Parallel code extraction with error handling
                code_contexts = {}
                extraction_errors = []
                
                def extract_context_safe(field):
                    try:
                        backend_xpath = field.get("backend_xpath") or ""
                        field_name = field.get("field_name", "")
                        
                        keywords = []
                        if backend_xpath:
                            last_seg = backend_xpath.split("/")[-1]
                            if last_seg:
                                keywords.append(last_seg)
                        if field_name:
                            keywords.append(field_name)
                        
                        if not keywords:
                            keywords = ["validate", "check"]
                        
                        snippets = extract_java_code_blocks_with_cross_references(
                            self.src_dir, keywords, max_depth=1
                        )
                        ctx = trim_code_context(snippets, max_chars=2000)
                        fid_short = f"{field_name}_{self._hash_text(backend_xpath or field_name)}"
                        return fid_short, ctx, None
                        
                    except Exception as e:
                        return None, "", str(e)
                
                # Execute code extraction in parallel
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    future_to_field = {executor.submit(extract_context_safe, field): field for field in batch}
                    
                    for future in as_completed(future_to_field):
                        field = future_to_field[future]
                        field_name = field.get('field_name', 'unknown')
                        
                        try:
                            fid_short, ctx, error = future.result()
                            if error:
                                extraction_errors.append(f"{field_name}: {error}")
                                # Continue with empty context
                                fid_short = f"{field_name}_{self._hash_text(field.get('backend_xpath', field_name))}"
                                ctx = ""
                            code_contexts[fid_short] = ctx
                        except Exception as e:
                            extraction_errors.append(f"{field_name}: {str(e)}")
                            fid_short = f"{field_name}_{self._hash_text(field.get('backend_xpath', field_name))}"
                            code_contexts[fid_short] = ""
                
                if extraction_errors:
                    print(f"[WARN] Code extraction errors in batch {batch_num}: {len(extraction_errors)} fields")
                
                # Sequential API call for the batch
                print(f"[INFO] Generating test cases for batch {batch_num}...")
                prompt = self._format_fiservai_prompt(batch, code_contexts)
                output = self._call_api_with_retry(prompt, max_retries=3)
                
                if output is None:
                    print(f"[ERROR] API call failed for batch {batch_num}")
                    batch_failures.append({
                        'batch': batch_num,
                        'fields': [f.get('field_name', 'unknown') for f in batch],
                        'error': 'API call failed'
                    })
                    total_failed += len(batch)
                    continue
                
                # Parse and store results
                success = self._parse_and_store(output, batch)
                if success:
                    total_processed += len(batch)
                    print(f"[INFO] Successfully processed batch {batch_num}")
                else:
                    print(f"[ERROR] Failed to parse results for batch {batch_num}")
                    batch_failures.append({
                        'batch': batch_num,
                        'fields': [f.get('field_name', 'unknown') for f in batch],
                        'error': 'Parsing failed'
                    })
                    total_failed += len(batch)
                
            except Exception as e:
                print(f"[ERROR] Batch {batch_num} failed with exception: {str(e)}")
                batch_failures.append({
                    'batch': batch_num,
                    'fields': [f.get('field_name', 'unknown') for f in batch],
                    'error': str(e)
                })
                total_failed += len(batch)
            
            # Brief pause between batches
            if i + batch_size < len(valid_fields):
                time.sleep(1)
        
        # Generate summary
        success_rate = (total_processed / len(valid_fields)) * 100 if valid_fields else 0
        
        summary = {
            'success': total_processed > 0,
            'processed': total_processed,
            'failed': total_failed,
            'success_rate': round(success_rate, 2),
            'invalid_fields': invalid_fields,
            'batch_failures': batch_failures,
            'total_test_cases': len(self.test_manager.get_all_cases())
        }
        
        print(f"\n[INFO] Bulk generation completed:")
        print(f"  - Processed: {total_processed}/{len(valid_fields)} fields ({success_rate:.1f}%)")
        print(f"  - Generated: {summary['total_test_cases']} test cases")
        print(f"  - Failed batches: {len(batch_failures)}")
        
        return summary

    def _parse_and_store(self, output: str, batch_fields: List[dict]) -> bool:
        """Parse AI output and add cases to TestCaseManager with error handling"""
        
        try:
            initial_count = len(self.test_manager.get_all_cases())
            
            for field in batch_fields:
                field_name = field.get('field_name', 'unknown')
                fid = field.get("backend_xpath") or field_name
                fid_short = f"{field_name}_{self._hash_text(fid)}"
                
                start_marker = f"===FIELD_START:{fid_short}==="
                end_marker = f"===FIELD_END:{fid_short}==="
                
                if start_marker in output and end_marker in output:
                    try:
                        block = output.split(start_marker, 1)[1].split(end_marker, 1)[0].strip()
                        self.test_manager.parse_and_add_test_cases(block, default_mapping=fid)
                        print(f"[DEBUG] Parsed test cases for field: {field_name}")
                    except Exception as e:
                        print(f"[WARN] Failed to parse marked section for {field_name}: {str(e)}")
                        # Try parsing the entire output as fallback
                        self.test_manager.parse_and_add_test_cases(output, default_mapping=fid)
                else:
                    print(f"[WARN] No field markers found for {field_name}, using entire output")
                    self.test_manager.parse_and_add_test_cases(output, default_mapping=fid)
            
            final_count = len(self.test_manager.get_all_cases())
            new_cases = final_count - initial_count
            
            if new_cases > 0:
                print(f"[INFO] Successfully added {new_cases} new test cases")
                return True
            else:
                print("[WARN] No new test cases were added")
                return False
                
        except Exception as e:
            print(f"[ERROR] Failed to parse and store test cases: {str(e)}")
            return False

    def get_generation_summary(self) -> Dict[str, Any]:
        """Get summary of generation results"""
        all_cases = self.test_manager.get_all_cases()
        
        return {
            'total_test_cases': len(all_cases),
            'failed_fields': self.failed_fields.copy(),
            'success_rate': 0 if not all_cases else 100 - (len(self.failed_fields) / (len(all_cases) + len(self.failed_fields)) * 100)
        }


# complete_agentic_test_generator.py (updated main functions)

import argparse
import os
from datetime import datetime
from typing import List, Dict, Any

from core.test_objective_core import TestObjectiveGeneratorCore
from core.testcase_manager import TestCaseManager
# from some_module import GPTClient, get_xpath_fields  # <-- keep your enterprise versions

def chat_mode_with_auto_export(generator: TestObjectiveGeneratorCore, fields: List[dict]) -> bool:
    """Chat mode with automatic export on exit"""
    
    print("ðŸ¤– CHAT MODE - Interactive Test Case Generation")
    print("Commands: 'list' - show available fields, 'exit' - quit and export")
    print("=" * 60)
    
    # Show available fields
    print(f"\nAvailable fields ({len(fields)}):")
    for i, field in enumerate(fields[:10]):  # Show first 10
        print(f"  {i+1}. {field.get('field_name', 'Unknown')}")
    if len(fields) > 10:
        print(f"  ... and {len(fields) - 10} more (type 'list' to see all)")
    
    generated_any = False
    
    while True:
        field_input = input(f"\nðŸ“ Enter field name (or 'list'/'exit'): ").strip()
        
        if field_input.lower() == "exit":
            break
            
        elif field_input.lower() == "list":
            print(f"\nAll available fields ({len(fields)}):")
            for i, field in enumerate(fields):
                print(f"  {i+1:2d}. {field.get('field_name', 'Unknown')}")
            continue
            
        elif field_input.lower() == "status":
            cases = generator.test_manager.get_all_cases()
            if cases:
                print(f"\nðŸ“Š Generated {len(cases)} test cases so far:")
                for case in cases[-5:]:  # Show last 5
                    print(f"  â€¢ {case.get('Test Case ID', 'N/A')}: {case.get('Test Objective', 'N/A')[:50]}...")
            else:
                print("\nðŸ“Š No test cases generated yet")
            continue
        
        # Find matching field (case-insensitive)
        field = None
        for f in fields:
            if f.get("field_name", "").lower() == field_input.lower():
                field = f
                break
        
        if not field:
            # Try partial match
            matches = [f for f in fields if field_input.lower() in f.get("field_name", "").lower()]
            if matches:
                if len(matches) == 1:
                    field = matches[0]
                    print(f"ðŸ” Found partial match: {field.get('field_name')}")
                else:
                    print(f"ðŸ” Multiple matches found:")
                    for i, match in enumerate(matches[:5]):
                        print(f"  {i+1}. {match.get('field_name')}")
                    continue
            else:
                print(f"âŒ Field '{field_input}' not found. Type 'list' to see available fields.")
                continue
        
        # Generate test cases for the field
        print(f"\nðŸ”„ Generating test cases for: {field.get('field_name')}")
        print("-" * 40)
        
        success = generator.generate_for_field(field)
        
        if success:
            print(f"âœ… Successfully generated test cases for {field.get('field_name')}")
            generated_any = True
            
            # Show what was generated
            all_cases = generator.test_manager.get_all_cases()
            if all_cases:
                recent_cases = all_cases[-3:]  # Show last 3 cases
                print(f"\nðŸ“‹ Recent test cases:")
                for case in recent_cases:
                    print(f"  â€¢ {case.get('Test Case ID', 'N/A')}: {case.get('Test Objective', 'N/A')}")
        else:
            print(f"âŒ Failed to generate test cases for {field.get('field_name')}")
    
    # Auto-export logic
    all_cases = generator.test_manager.get_all_cases()
    
    if not all_cases:
        print("\nðŸ¤· No test cases were generated during this session.")
        return False
    
    print(f"\nðŸ“¤ CHAT SESSION COMPLETE")
    print(f"Generated {len(all_cases)} test cases total")
    
    # Show summary of what was generated
    print(f"\nðŸ“‹ Generated test cases:")
    for i, case in enumerate(all_cases, 1):
        print(f"  {i:2d}. {case.get('Test Case ID', 'N/A')}: {case.get('Test Objective', 'N/A')[:60]}...")
    
    # Auto-export prompt
    print(f"\nðŸ’¾ AUTOMATIC EXPORT")
    export_choice = input("Save all test cases to Excel? (Y/n): ").strip().lower()
    
    if export_choice in ['', 'y', 'yes']:
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"chat_mode_test_cases_{timestamp}.xlsx"
            filepath = os.path.join(os.getcwd(), filename)
            
            generator.test_manager.export_to_excel(filepath)
            print(f"âœ… Test cases exported to: {filename}")
            return True
            
        except Exception as e:
            print(f"âŒ Export failed: {str(e)}")
            return False
    else:
        print("âŒ Export cancelled. Test cases remain in memory only.")
        return False

def bulk_mode_with_error_handling(generator: TestObjectiveGeneratorCore, fields: List[dict], out_file: str) -> bool:
    """Bulk mode with comprehensive error handling and reporting"""
    
    print("ðŸš€ BULK MODE - Processing All Fields")
    print("=" * 50)
    
    # Show what will be processed
    print(f"ðŸ“Š Processing {len(fields)} fields")
    print(f"ðŸ’¾ Output file: {out_file}")
    
    # Confirm before starting
    proceed = input("\nðŸš¦ Proceed with bulk generation? (Y/n): ").strip().lower()
    if proceed not in ['', 'y', 'yes']:
        print("âŒ Bulk generation cancelled")
        return False
    
    # Start bulk generation
    print(f"\nðŸ”„ Starting bulk generation...")
    start_time = datetime.now()
    
    summary = generator.bulk_generate(fields, batch_size=5)
    
    end_time = datetime.now()
    duration = end_time - start_time
    
    # Show detailed results
    print(f"\nðŸ“ˆ BULK GENERATION RESULTS")
    print("=" * 50)
    print(f"â±ï¸  Duration: {duration}")
    print(f"âœ… Successful: {summary['processed']} fields")
    print(f"âŒ Failed: {summary['failed']} fields") 
    print(f"ðŸ“Š Success Rate: {summary['success_rate']}%")
    print(f"ðŸ“‹ Total Test Cases: {summary['total_test_cases']}")
    
    # Show failures if any
    if summary['batch_failures']:
        print(f"\nâŒ BATCH FAILURES ({len(summary['batch_failures'])}):")
        for failure in summary['batch_failures']:
            print(f"  Batch {failure['batch']}: {failure['error']}")
            print(f"    Fields: {', '.join(failure['fields'][:3])}{'...' if len(failure['fields']) > 3 else ''}")
    
    if summary['invalid_fields']:
        print(f"\nâš ï¸  INVALID FIELDS ({len(summary['invalid_fields'])}):")
        for invalid in summary['invalid_fields'][:5]:
            print(f"  â€¢ {invalid['field']}: {invalid['error']}")
        if len(summary['invalid_fields']) > 5:
            print(f"  ... and {len(summary['invalid_fields']) - 5} more")
    
    # Export results
    if summary['total_test_cases'] > 0:
        try:
            print(f"\nðŸ’¾ Exporting {summary['total_test_cases']} test cases to Excel...")
            generator.test_manager.export_to_excel(out_file)
            print(f"âœ… Export successful: {out_file}")
            
            # Create summary report
            summary_file = out_file.replace('.xlsx', '_summary.txt')
            with open(summary_file, 'w') as f:
                f.write(f"Bulk Generation Summary - {datetime.now()}\n")
                f.write("=" * 50 + "\n")
                f.write(f"Duration: {duration}\n")
                f.write(f"Fields Processed: {summary['processed']}/{len(fields)}\n")
                f.write(f"Success Rate: {summary['success_rate']}%\n")
                f.write(f"Test Cases Generated: {summary['total_test_cases']}\n\n")
                
                if summary['batch_failures']:
                    f.write("Batch Failures:\n")
                    for failure in summary['batch_failures']:
                        f.write(f"  Batch {failure['batch']}: {failure['error']}\n")
                        f.write(f"    Fields: {', '.join(failure['fields'])}\n")
                
            print(f"ðŸ“„ Summary report: {summary_file}")
            return True
            
        except Exception as e:
            print(f"âŒ Export failed: {str(e)}")
            return False
    else:
        print("âŒ No test cases generated - nothing to export")
        return False

def main():
    parser = argparse.ArgumentParser(description="Agentic Test Case Generator")
    parser.add_argument("--mode", choices=["chat", "bulk"], help="Run in chatbot or bulk mode")
    parser.add_argument("--mapping", help="Path to mapping sheet")
    parser.add_argument("--src", help="Path to Java source code")
    parser.add_argument("--out", default="test_objectives.xlsx", help="Output Excel file (bulk mode)")
    
    args = parser.parse_args()
    
    # Interactive prompts if args missing
    if not args.mode:
        print("ðŸ¤– Select Generation Mode:")
        print("  1. Bulk Mode (process all fields â†’ Excel)")
        print("  2. Chat Mode (interactive field selection)")
        mode_input = input("Choose (1/2): ").strip()
        args.mode = "bulk" if mode_input == "1" else "chat"
    
    if not args.mapping:
        args.mapping = input("ðŸ“Š Enter path to mapping file: ").strip()
    
    if not args.src:
        args.src = input("ðŸ“ Enter path to Java source directory: ").strip()
    
    # Validate inputs
    if not os.path.exists(args.mapping):
        print(f"âŒ Mapping file not found: {args.mapping}")
        return
    
    if not os.path.isdir(args.src):
        print(f"âŒ Source directory not found: {args.src}")
        return
    
    try:
        # Load fields (replace with your actual implementation)
        # fields = get_xpath_fields(args.mapping, with_metadata=True)
        fields = []  # Placeholder - use your actual field loading
        
        if not fields:
            print("âŒ No fields found in mapping file")
            return
        
        print(f"âœ… Loaded {len(fields)} fields from mapping")
        
        # Initialize components (replace with your actual implementations)
        # client = GPTClient()  # Your enterprise GPT client
        client = None  # Placeholder
        
        manager = TestCaseManager()
        generator = TestObjectiveGeneratorCore(client, manager, args.src)
        
        # Run selected mode
        if args.mode == "chat":
            success = chat_mode_with_auto_export(generator, fields)
        else:
            success = bulk_mode_with_error_handling(generator, fields, args.out)
        
        if success:
            print(f"\nðŸŽ‰ Generation completed successfully!")
        else:
            print(f"\nâš ï¸  Generation completed with issues. Check the logs above.")
    
    except Exception as e:
        print(f"âŒ Fatal error: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()


# core/testcase_manager.py (enhanced version)

import pandas as pd
import os
from typing import List, Dict, Any, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class TestCaseManager:
    def __init__(self):
        self.test_cases: List[Dict[str, Any]] = []
        self.counter = 1
        self.parse_errors = []  # Track parsing errors
        
    def parse_and_add_test_cases(self, raw_text: str, default_mapping: str = "") -> int:
        """Parse raw AI output (tab-separated rows) and add to internal store.
        Returns number of cases added."""
        
        if not raw_text or not raw_text.strip():
            print("[WARN] Empty input provided for parsing")
            return 0
        
        initial_count = len(self.test_cases)
        lines = raw_text.strip().splitlines()
        
        for line_num, line in enumerate(lines, 1):
            line = line.strip()
            
            # Skip empty lines, headers, and obvious non-data lines
            if (not line or 
                line.startswith('Category') or 
                line.startswith('---') or 
                line.startswith('===') or
                'Test Case ID' in line or
                line.startswith('Sorry') or
                line.startswith('I don\'t') or
                len(line) < 10):  # Too short to be a real test case
                continue
            
            try:
                # Try tab separation first
                parts = line.split('\t')
                
                # Fallback to pipe separation if not enough parts
                if len(parts) < 6:
                    parts = line.split('|')
                
                # Still not enough parts - log and skip
                if len(parts) < 6:
                    error_msg = f"Line {line_num}: Insufficient columns ({len(parts)}) - '{line[:50]}...'"
                    self.parse_errors.append(error_msg)
                    print(f"[WARN] {error_msg}")
                    continue
                
                # Clean and pad parts
                parts = [p.strip() for p in parts]
                while len(parts) < 9:
                    parts.append("")
                
                # Extract components with validation
                category = parts[0] or "Functional"
                tc_id = parts[1].strip()
                val_type = parts[2] or "Field Validation - Positive" 
                objective = parts[3]
                req_field = parts[4] or "Request"
                steps = parts[5]
                expected = parts[6]
                mapping = parts[7] or default_mapping
                mode = parts[8]
                
                # Validate required fields
                if not objective or len(objective.strip()) < 5:
                    error_msg = f"Line {line_num}: Missing or invalid test objective"
                    self.parse_errors.append(error_msg)
                    print(f"[WARN] {error_msg}")
                    continue
                
                if not steps or len(steps.strip()) < 5:
                    error_msg = f"Line {line_num}: Missing or invalid test steps"
                    self.parse_errors.append(error_msg)
                    print(f"[WARN] {error_msg}")
                    continue
                
                # Auto-generate ID if missing or invalid
                if not tc_id or not tc_id.startswith('TC_'):
                    tc_id = f"TC_{self.counter:03d}"
                    self.counter += 1
                
                # Validate and fix validation type
                valid_types = [
                    "Field Validation - Positive",
                    "Field Validation - Negative", 
                    "Business Validation - Positive",
                    "Business Validation - Negative"
                ]
                
                if val_type not in valid_types:
                    # Try to map common variations
                    val_type_lower = val_type.lower()
                    if "positive" in val_type_lower and "field" in val_type_lower:
                        val_type = "Field Validation - Positive"
                    elif "negative" in val_type_lower and "field" in val_type_lower:
                        val_type = "Field Validation - Negative"
                    elif "positive" in val_type_lower and "business" in val_type_lower:
                        val_type = "Business Validation - Positive"
                    elif "negative" in val_type_lower and "business" in val_type_lower:
                        val_type = "Business Validation - Negative"
                    else:
                        val_type = "Field Validation - Positive"  # Default
                
                # Auto-set manual/automation based on validation type
                if not mode:
                    if "Business" in val_type:
                        mode = "Manual"
                    else:
                        mode = "Automation"
                
                # Create test case
                case = {
                    "Category": category,
                    "Test Case ID": tc_id,
                    "Type of Validation": val_type,
                    "Test Objective": objective,
                    "Request/Response Field": req_field,
                    "Test Steps": steps,
                    "Expected Result": expected,
                    "Mapping Correlation": mapping,
                    "Manual/Automation": mode
                }
                
                self.test_cases.append(case)
                
                # Update counter to avoid duplicates
                if tc_id.startswith('TC_'):
                    try:
                        tc_num = int(tc_id.split('_')[1])
                        self.counter = max(self.counter, tc_num + 1)
                    except (IndexError, ValueError):
                        pass
                
            except Exception as e:
                error_msg = f"Line {line_num}: Parse error - {str(e)}"
                self.parse_errors.append(error_msg)
                print(f"[ERROR] {error_msg}")
                continue
        
        added_count = len(self.test_cases) - initial_count
        
        if added_count > 0:
            print(f"[INFO] Successfully parsed {added_count} test cases")
        else:
            print("[WARN] No valid test cases found in input")
            
        return added_count
    
    def get_all_cases(self) -> List[Dict[str, Any]]:
        """Get all stored test cases"""
        return self.test_cases.copy()
    
    def get_parse_errors(self) -> List[str]:
        """Get all parsing errors encountered"""
        return self.parse_errors.copy()
    
    def clear_cases(self):
        """Clear all test cases and reset counter"""
        self.test_cases.clear()
        self.counter = 1
        self.parse_errors.clear()
        print("[INFO] All test cases cleared")
    
    def export_to_excel(self, out_file: str) -> bool:
        """Export all stored cases to Excel with professional formatting"""
        
        if not self.test_cases:
            print("[ERROR] No test cases to export")
            return False
        
        try:
            # Create DataFrame
            df = pd.DataFrame(self.test_cases)
            
            # Ensure all required columns are present
            required_columns = [
                "Category", "Test Case ID", "Type of Validation", 
                "Test Objective", "Request/Response Field", "Test Steps", 
                "Expected Result", "Mapping Correlation", "Manual/Automation"
            ]
            
            for col in required_columns:
                if col not in df.columns:
                    df[col] = ""
            
            # Reorder columns
            df = df[required_columns]
            
            # Create Excel with formatting
            self._create_formatted_excel(df, out_file)
            
            print(f"[INFO] Exported {len(self.test_cases)} test cases to {out_file}")
            return True
            
        except Exception as e:
            print(f"[ERROR] Export failed: {str(e)}")
            return False
    
    def _create_formatted_excel(self, df: pd.DataFrame, filepath: str):
        """Create Excel with professional formatting"""
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Define styles
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='Calibri', size=10)
        data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        
        # Add headers with formatting
        for col, header in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Add data with formatting
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                
                # Color coding for validation types
                if col_idx == 3:  # Type of Validation column
                    if 'Positive' in str(value):
                        cell.fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
                    elif 'Negative' in str(value):
                        cell.fill = PatternFill(start_color='FFF0F0', end_color='FFF0F0', fill_type='solid')
                
                # Highlight manual tests
                if col_idx == 9 and str(value).lower() == 'manual':  # Manual/Automation column
                    cell.fill = PatternFill(start_color='FFF8DC', end_color='FFF8DC', fill_type='solid')
        
        # Auto-adjust column widths
        column_widths = [12, 15, 25, 45, 18, 45, 35, 30, 15]
        for col, width in enumerate(column_widths, 1):
            if col <= len(df.columns):
                column_letter = ws.cell(row=1, column=col).column_letter
                ws.column_dimensions[column_letter].width = width
        
        # Set row heights
        ws.row_dimensions[1].height = 35  # Header row
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_data = self._generate_summary_data()
        
        summary_ws.append(["Test Case Generation Summary"])
        summary_ws.append([])
        
        for key, value in summary_data.items():
            summary_ws.append([key, value])
        
        # Format summary sheet
        summary_ws.cell(1, 1).font = Font(size=14, bold=True)
        for row in range(1, summary_ws.max_row + 1):
            for col in range(1, 3):
                cell = summary_ws.cell(row, col)
                cell.border = thin_border
                if row > 2:
                    cell.font = Font(size=10)
        
        summary_ws.column_dimensions['A'].width = 25
        summary_ws.column_dimensions['B'].width = 15
        
        # Save workbook
        wb.save(filepath)
    
    def _generate_summary_data(self) -> Dict[str, Any]:
        """Generate summary statistics"""
        
        if not self.test_cases:
            return {"Total Test Cases": 0}
        
        total = len(self.test_cases)
        
        # Count by validation type
        validation_counts = {}
        manual_count = 0
        automation_count = 0
        
        for case in self.test_cases:
            val_type = case.get('Type of Validation', 'Unknown')
            validation_counts[val_type] = validation_counts.get(val_type, 0) + 1
            
            mode = case.get('Manual/Automation', '').lower()
            if mode == 'manual':
                manual_count += 1
            elif mode == 'automation':
                automation_count += 1
        
        summary = {
            "Generation Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Total Test Cases": total,
            "Manual Tests": manual_count,
            "Automation Tests": automation_count,
            "Parse Errors": len(self.parse_errors)
        }
        
        # Add validation type breakdown
        for val_type, count in validation_counts.items():
            summary[f"{val_type} Tests"] = count
        
        return summary
    
    def get_statistics(self) -> Dict[str, Any]:
        """Get current statistics"""
        return self._generate_summary_data()
