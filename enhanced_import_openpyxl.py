import openpyxl
from typing import List, Dict, Set, Union, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed
import os

# Global variables to store file paths
mapping_file = "PartyInq-PRM10.0_v11.0_R2025.2.xlsm"
yaml_file = ""
output_yaml_file = ""

def get_xpath_fields(file_path: str, with_metadata: bool = False, target_xpath: Optional[str] = None) -> Union[List[Dict], Set[str]]:
    """
    Extract field information from Excel mapping file.
    
    Args:
        file_path: Path to Excel file
        with_metadata: If True, returns full metadata; if False, returns just xpath names
        target_xpath: If specified, returns metadata only for this specific field
    
    Returns:
        - If target_xpath specified: List with single field dict or empty list
        - If with_metadata=True: List of field metadata dicts
        - If with_metadata=False: Set of xpath strings
    """
    print(f"[INFO] Opening file: {file_path}")
    
    # Validate file exists
    if not os.path.exists(file_path):
        print(f"[ERROR] File not found: {file_path}")
        return [] if with_metadata else set()
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"[ERROR] Error opening file: {e}")
        return [] if with_metadata else set()

    if "EFX Object Mapping" not in wb.sheetnames:
        print("[ERROR] Sheet 'EFX Object Mapping' not found.")
        print(f"[INFO] Available sheets: {wb.sheetnames}")
        return [] if with_metadata else set()

    sheet = wb["EFX Object Mapping"]
    print("[INFO] Sheet loaded successfully.")

    # If target xpath is set, only return that field's metadata (optimized for chat mode)
    if target_xpath:
        print(f"[INFO] Searching for specific field: {target_xpath}")
        
        for row_tuple in sheet.iter_rows(min_row=3, max_row=sheet.max_row, values_only=False):
            xpath_cell = row_tuple[1]  # Column B (XPATH)
            
            if xpath_cell.value and str(xpath_cell.value).strip() == target_xpath:
                field_metadata = _extract_field_metadata(row_tuple, sheet)
                print("[INFO] Extraction complete. 1 field found.")
                return [field_metadata] if with_metadata else {xpath_cell.value}
        
        print(f"[WARN] Field '{target_xpath}' not found in mapping sheet.")
        return [] if with_metadata else set()

    # Default: extract all fields (for bulk mode or field list)
    results = [] if with_metadata else set()
    processed_count = 0
    skipped_count = 0
    
    print("[INFO] Processing all fields...")
    
    for row_tuple in sheet.iter_rows(min_row=3, max_row=sheet.max_row, values_only=False):
        xpath_cell = row_tuple[1]  # Column B (XPATH)
        datatype_cell = row_tuple[4]  # Column E (DataType)
        
        # Skip hidden rows
        row_number = xpath_cell.row
        if sheet.row_dimensions.get(row_number) and sheet.row_dimensions[row_number].hidden:
            continue

        # Skip rows where DataType is "aggregate"
        if datatype_cell.value and str(datatype_cell.value).strip().lower() == "aggregate":
            skipped_count += 1
            continue

        value = xpath_cell.value
        if value and str(value).strip():  # Non-empty xpath
            if with_metadata:
                field_metadata = _extract_field_metadata(row_tuple, sheet)
                results.append(field_metadata)
            else:
                results.add(str(value).strip())
            
            processed_count += 1

    print(f"[INFO] Extraction complete. Processed: {processed_count}, Skipped: {skipped_count}")
    return results

def _extract_field_metadata(row_tuple, sheet) -> Dict[str, str]:
    """Extract metadata from a single row tuple."""
    
    xpath_cell = row_tuple[1]           # Column B (XPATH)
    description_cell = row_tuple[2]     # Column C (Description)  
    usage_cell = row_tuple[3]           # Column D (Usage)
    datatype_cell = row_tuple[4]        # Column E (DataType)
    efx_allowed_values_cell = row_tuple[7]  # Column H (EFX Allowed Values)
    service_notes_cell = row_tuple[8]   # Column I (Service Provider Implementation Notes)
    published_description_cell = row_tuple[9]  # Column J (Published Description)
    backend_xpath_cell = row_tuple[10]  # Column K (Backend Acceptable XPATH)
    backend_datatype_cell = row_tuple[12]  # Column M (Backend Acceptable DataType)
    transformation_rules_cell = row_tuple[14]  # Column O (Transformation Rules)
    
    # Helper function to safely get cell value
    def safe_cell_value(cell):
        return str(cell.value).strip() if cell.value else ""
    
    field_metadata = {
        "field_name": safe_cell_value(xpath_cell).split('/')[-1] if safe_cell_value(xpath_cell) else "",
        "xpath": safe_cell_value(xpath_cell),
        "description": safe_cell_value(description_cell),
        "usage": safe_cell_value(usage_cell),
        "datatype": safe_cell_value(datatype_cell),
        "efx_allowed_values": safe_cell_value(efx_allowed_values_cell),
        "service_notes": safe_cell_value(service_notes_cell),
        "published_description": safe_cell_value(published_description_cell),
        "backend_xpath": safe_cell_value(backend_xpath_cell),
        "backend_datatype": safe_cell_value(backend_datatype_cell),
        "transformation_rules": safe_cell_value(transformation_rules_cell),
    }
    
    return field_metadata

def get_field_names_only(file_path: str) -> Set[str]:
    """
    Optimized function to get only field names (for chat mode field selection).
    Much faster than loading full metadata.
    """
    return get_xpath_fields(file_path, with_metadata=False)

def get_field_metadata_batch(file_path: str, field_names: List[str], max_workers: int = 4) -> List[Dict[str, str]]:
    """
    Get metadata for multiple specific fields in parallel (for batch processing).
    
    Args:
        file_path: Path to Excel file
        field_names: List of xpath field names to extract
        max_workers: Maximum number of parallel workers
    
    Returns:
        List of field metadata dicts
    """
    print(f"[INFO] Loading metadata for {len(field_names)} fields in parallel...")
    
    results = []
    failed_fields = []
    
    def load_single_field(field_name):
        try:
            field_data = get_xpath_fields(file_path, with_metadata=True, target_xpath=field_name)
            return field_data[0] if field_data else None
        except Exception as e:
            print(f"[ERROR] Failed to load field {field_name}: {e}")
            return None
    
    # Use ThreadPoolExecutor for parallel loading
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_field = {executor.submit(load_single_field, field): field for field in field_names}
        
        for future in as_completed(future_to_field):
            field_name = future_to_field[future]
            try:
                result = future.result()
                if result:
                    results.append(result)
                else:
                    failed_fields.append(field_name)
            except Exception as e:
                print(f"[ERROR] Exception loading field {field_name}: {e}")
                failed_fields.append(field_name)
    
    if failed_fields:
        print(f"[WARN] Failed to load {len(failed_fields)} fields: {failed_fields[:3]}{'...' if len(failed_fields) > 3 else ''}")
    
    print(f"[INFO] Successfully loaded metadata for {len(results)}/{len(field_names)} fields")
    return results

def validate_mapping_file(file_path: str) -> bool:
    """Validate that the mapping file exists and has the correct structure."""
    
    if not os.path.exists(file_path):
        print(f"[ERROR] Mapping file not found: {file_path}")
        return False
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if "EFX Object Mapping" not in wb.sheetnames:
            print(f"[ERROR] Required sheet 'EFX Object Mapping' not found")
            print(f"[INFO] Available sheets: {wb.sheetnames}")
            return False
        
        # Check if sheet has data
        sheet = wb["EFX Object Mapping"]
        if sheet.max_row < 3:
            print(f"[ERROR] Sheet appears to be empty (max_row: {sheet.max_row})")
            return False
        
        print(f"[INFO] Mapping file validated successfully")
        return True
        
    except Exception as e:
        print(f"[ERROR] Error validating mapping file: {e}")
        return False

# Updated main function for testing
def main_extract_hardcoded_field():
    """Test function - can be removed in production"""
    global mapping_file
    
    hardcoded_field = "PartyRec/PersonPartyInfo/PersonData/Contact/PostAddr/PostalCode"
    
    if not validate_mapping_file(mapping_file):
        return
    
    # Test single field extraction
    fields_metadata = get_xpath_fields(mapping_file, with_metadata=True, target_xpath=hardcoded_field)
    
    if fields_metadata:
        meta = fields_metadata[0]
        print(f"[INFO] Metadata for {hardcoded_field}:")
        for k, v in meta.items():
            print(f"  {k}: {v}")
    else:
        print(f"[ERROR] Field '{hardcoded_field}' not found in mapping sheet.")

# Integration helpers for the agentic system
class FieldMetadataLoader:
    """Helper class for integrating with the agentic test generator"""
    
    def __init__(self, mapping_file_path: str):
        self.mapping_file_path = mapping_file_path
        self._field_names_cache = None
    
    def get_available_fields(self) -> Set[str]:
        """Get list of all available field names (cached for performance)"""
        if self._field_names_cache is None:
            self._field_names_cache = get_field_names_only(self.mapping_file_path)
        return self._field_names_cache
    
    def get_field_metadata(self, field_name: str) -> Optional[Dict[str, str]]:
        """Get metadata for a single field (for chat mode)"""
        results = get_xpath_fields(self.mapping_file_path, with_metadata=True, target_xpath=field_name)
        return results[0] if results else None
    
    def get_all_field_metadata(self) -> List[Dict[str, str]]:
        """Get metadata for all fields (for bulk mode)"""
        return get_xpath_fields(self.mapping_file_path, with_metadata=True)
    
    def get_batch_metadata(self, field_names: List[str]) -> List[Dict[str, str]]:
        """Get metadata for specific fields in parallel"""
        return get_field_metadata_batch(self.mapping_file_path, field_names)
    
    def validate_file(self) -> bool:
        """Validate the mapping file"""
        return validate_mapping_file(self.mapping_file_path)

if __name__ == "__main__":
    main_extract_hardcoded_field()
